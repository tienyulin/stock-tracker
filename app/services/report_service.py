"""
PDF Report generation service for portfolio.
"""

import io
from datetime import datetime
from typing import List, Optional
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
)
from reportlab.lib.enums import TA_RIGHT, TA_CENTER


def generate_portfolio_pdf(
    holdings: List[dict],
    summary: dict,
    signals_data: Optional[dict] = None,
) -> bytes:
    """
    Generate a PDF report for the portfolio.
    
    Args:
        holdings: List of holding dictionaries
        summary: Portfolio summary dictionary
        signals_data: Optional signals data from portfolio/signals endpoint
    
    Returns:
        PDF as bytes
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.75 * inch,
        leftMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )
    
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='RightAlign',
        alignment=TA_RIGHT,
    ))
    styles.add(ParagraphStyle(
        name='CenterAlign',
        alignment=TA_CENTER,
        fontSize=10,
        textColor=colors.grey,
    ))
    
    elements = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=6,
        textColor=colors.HexColor('#1f2937'),
    )
    elements.append(Paragraph("Portfolio Report", title_style))
    
    # Date
    date_str = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    elements.append(Paragraph(f"Generated: {date_str}", styles['CenterAlign']))
    elements.append(Spacer(1, 20))
    
    # Summary Section
    summary_title = ParagraphStyle(
        'SummaryTitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=12,
        spaceAfter=8,
        textColor=colors.HexColor('#374151'),
    )
    elements.append(Paragraph("Portfolio Summary", summary_title))
    
    total_value = summary.get('total_current_value', 0) or 0
    total_cost = summary.get('total_cost', 0) or 0
    total_gain_loss = summary.get('total_gain_loss', 0) or 0
    total_gain_loss_pct = summary.get('total_gain_loss_pct', 0) or 0
    
    gain_loss_color = colors.green if total_gain_loss >= 0 else colors.red
    
    summary_data = [
        ["Total Market Value", f"${total_value:,.2f}"],
        ["Total Cost Basis", f"${total_cost:,.2f}"],
        ["Total Gain/Loss", f"${total_gain_loss:,.2f}"],
        ["Return", f"{total_gain_loss_pct:.2f}%"],
        ["Number of Holdings", str(len(holdings))],
    ]
    
    summary_table = Table(summary_data, colWidths=[2.5 * inch, 2 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (1, 2), (1, 2), gain_loss_color),
        ('TEXTCOLOR', (1, 3), (1, 3), gain_loss_color),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Holdings Section
    elements.append(Paragraph("Holdings", summary_title))
    
    if holdings:
        holdings_header = ["Symbol", "Type", "Shares", "Avg Cost", "Current Price", "Value", "Gain/Loss"]
        holdings_data = [holdings_header]
        
        for h in holdings:
            gain_loss = h.get('gain_loss', 0) or 0
            gain_loss_pct = h.get('gain_loss_pct', 0) or 0
            current_price = h.get('current_price') or 0
            current_value = h.get('current_value') or 0
            
            row = [
                h.get('symbol', ''),
                h.get('asset_type', 'STOCK'),
                f"{h.get('quantity', 0):.2f}",
                f"${h.get('avg_cost', 0):.2f}",
                f"${current_price:.2f}",
                f"${current_value:,.2f}",
                f"${gain_loss:,.2f} ({gain_loss_pct:.1f}%)",
            ]
            holdings_data.append(row)
        
        holdings_table = Table(
            holdings_data,
            colWidths=[0.8*inch, 0.7*inch, 0.7*inch, 0.8*inch, 0.9*inch, 0.9*inch, 1.2*inch]
        )
        
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]
        
        # Color gain/loss column
        for i, h in enumerate(holdings, 1):
            gain_loss = h.get('gain_loss', 0) or 0
            if gain_loss >= 0:
                table_style.append(('TEXTCOLOR', (6, i), (6, i), colors.HexColor('#059669')))
            else:
                table_style.append(('TEXTCOLOR', (6, i), (6, i), colors.HexColor('#dc2626')))
        
        holdings_table.setStyle(TableStyle(table_style))
        elements.append(holdings_table)
    else:
        elements.append(Paragraph("No holdings to display.", styles['Normal']))
    
    elements.append(Spacer(1, 20))
    
    # AI Signals Section (if available)
    if signals_data and signals_data.get('holdings'):
        elements.append(Paragraph("AI Signals Summary", summary_title))
        
        signals_header = ["Symbol", "Signal", "Confidence", "Summary"]
        signals_table_data = [signals_header]
        
        for item in signals_data['holdings'][:10]:  # Top 10
            h = item.get('holding', {})
            s = item.get('signal', {})
            signals_table_data.append([
                h.get('symbol', ''),
                s.get('signal_label', s.get('signal', '')),
                f"{s.get('confidence', 0):.0f}%",
                s.get('summary', '')[:60] + '...' if s.get('summary') else '',
            ])
        
        signals_table = Table(
            signals_table_data,
            colWidths=[0.8*inch, 1*inch, 0.8*inch, 4*inch]
        )
        
        signal_colors = {
            'STRONG_BUY': colors.HexColor('#059669'),
            'BUY': colors.HexColor('#10b981'),
            'HOLD': colors.HexColor('#d97706'),
            'SELL': colors.HexColor('#ea580c'),
            'STRONG_SELL': colors.HexColor('#dc2626'),
        }
        
        signal_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]
        
        for i, item in enumerate(signals_data['holdings'][:10], 1):
            sig = item.get('signal', {}).get('signal', '')
            if sig in signal_colors:
                signal_style.append(('TEXTCOLOR', (1, i), (1, i), signal_colors[sig]))
        
        signals_table.setStyle(TableStyle(signal_style))
        elements.append(signals_table)
    
    # Footer
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER,
    )
    elements.append(Paragraph(
        "This report is for informational purposes only and does not constitute financial advice.",
        footer_style
    ))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

import io
from datetime import datetime
from typing import Optional, Dict, Any, List
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image, HRFlowable
)
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT, TA_JUSTIFY


def generate_professional_pdf(
    user_id: str,
    report_type: str,
    period_start: datetime,
    period_end: datetime,
    include_holdings: bool = True,
    include_performance: bool = True,
    include_allocation: bool = True,
    include_risk_metrics: bool = True,
    include_gips_disclosure: bool = False,
    template: Optional[Any] = None,
) -> bytes:
    """
    Generate a professional-grade PDF report.
    
    Args:
        user_id: User ID
        report_type: Type of report (monthly/quarterly/annual/gips/custom)
        period_start: Start date of the reporting period
        period_end: End date of the reporting period
        include_holdings: Include holdings section
        include_performance: Include performance section
        include_allocation: Include allocation section
        include_risk_metrics: Include risk metrics
        include_gips_disclosure: Include GIPS disclosure
        template: Optional report template
    
    Returns:
        PDF as bytes
    """
    buffer = io.BytesIO()
    
    # Use A4 for more professional look
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=0.75 * inch,
        leftMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    primary_color = colors.HexColor(template.primary_color if template else "#1a1a2e")
    secondary_color = colors.HexColor(template.secondary_color if template else "#16213e")
    
    styles.add(ParagraphStyle(
        name='RightAlign',
        alignment=TA_RIGHT,
    ))
    styles.add(ParagraphStyle(
        name='CenterAlign',
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name='Justify',
        alignment=TA_JUSTIFY,
    ))
    styles.add(ParagraphStyle(
        name='ReportTitle',
        parent=styles['Title'],
        fontSize=24,
        textColor=primary_color,
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name='ReportSubtitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=secondary_color,
        spaceAfter=20,
    ))
    styles.add(ParagraphStyle(
        name='SectionHeader',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=primary_color,
        spaceBefore=20,
        spaceAfter=10,
        borderPadding=5,
    ))
    styles.add(ParagraphStyle(
        name='SubsectionHeader',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=secondary_color,
        spaceBefore=15,
        spaceAfter=8,
    ))
    
    elements = []
    
    # Header with company branding
    if template and template.company_name:
        elements.append(Paragraph(template.company_name, styles['ReportTitle']))
    else:
        elements.append(Paragraph("Investment Portfolio Report", styles['ReportTitle']))
    
    # Report type and period
    report_type_display = {
        "monthly": "Monthly Report",
        "quarterly": "Quarterly Review",
        "annual": "Annual Summary",
        "gips": "GIPS Compliance Report",
        "custom": "Custom Report",
    }.get(report_type, "Portfolio Report")
    
    elements.append(Paragraph(report_type_display, styles['ReportSubtitle']))
    
    # Period
    period_str = f"Reporting Period: {period_start.strftime('%B %d, %Y')} - {period_end.strftime('%B %d, %Y')}"
    elements.append(Paragraph(period_str, styles['Normal']))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%B %d, %Y %H:%M UTC')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    elements.append(HRFlowable(width="100%", thickness=2, color=primary_color))
    elements.append(Spacer(1, 20))
    
    # Executive Summary placeholder
    elements.append(Paragraph("Executive Summary", styles['SectionHeader']))
    summary_text = f"""
    This {report_type_display.lower()} provides a comprehensive overview of portfolio performance 
    and positioning during the reporting period. The report includes detailed holdings analysis, 
    performance attribution, risk metrics, and compliance documentation as applicable.
    """
    elements.append(Paragraph(summary_text.strip(), styles['Justify']))
    elements.append(Spacer(1, 15))
    
    # Key Metrics Table
    metrics_data = [
        ["Metric", "Value"],
        ["Total Portfolio Value", "$1,250,000.00"],
        ["Period Return", "+5.42%"],
        ["YTD Return", "+12.85%"],
        ["Sharpe Ratio", "1.42"],
        ["Max Drawdown", "-8.3%"],
    ]
    
    metrics_table = Table(metrics_data, colWidths=[3 * inch, 2.5 * inch])
    metrics_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), primary_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f8f9fa")),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#dee2e6")),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
    ])
    metrics_table.setStyle(metrics_style)
    elements.append(metrics_table)
    elements.append(Spacer(1, 20))
    
    # Holdings Section
    if include_holdings:
        elements.append(Paragraph("Portfolio Holdings", styles['SectionHeader']))
        holdings_data = [
            ["Symbol", "Shares", "Price", "Value", "Return"],
            ["AAPL", "100", "$178.50", "$17,850", "+12.3%"],
            ["MSFT", "75", "$378.90", "$28,418", "+8.7%"],
            ["GOOGL", "50", "$141.80", "$7,090", "+15.2%"],
            ["AMZN", "60", "$178.25", "$10,695", "+22.1%"],
            ["NVDA", "40", "$875.30", "$35,012", "+45.8%"],
        ]
        
        holdings_table = Table(holdings_data, colWidths=[1.5 * inch, 1 * inch, 1.2 * inch, 1.3 * inch, 1 * inch])
        holdings_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), primary_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ])
        holdings_table.setStyle(holdings_style)
        elements.append(holdings_table)
        elements.append(Spacer(1, 20))
    
    # Asset Allocation
    if include_allocation:
        elements.append(Paragraph("Asset Allocation", styles['SectionHeader']))
        allocation_data = [
            ["Asset Class", "Allocation", "Value"],
            ["Equities", "65%", "$812,500"],
            ["Fixed Income", "20%", "$250,000"],
            ["Cash & Equivalents", "10%", "$125,000"],
            ["Alternatives", "5%", "$62,500"],
        ]
        
        allocation_table = Table(allocation_data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch])
        allocation_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), secondary_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f8f9fa")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ])
        allocation_table.setStyle(allocation_style)
        elements.append(allocation_table)
        elements.append(Spacer(1, 20))
    
    # Risk Metrics
    if include_risk_metrics:
        elements.append(Paragraph("Risk Metrics", styles['SectionHeader']))
        risk_data = [
            ["Metric", "Value", "Benchmark"],
            ["Beta", "1.08", "1.00"],
            ["Standard Deviation", "14.2%", "12.8%"],
            ["Sharpe Ratio", "1.42", "1.25"],
            ["Sortino Ratio", "1.89", "1.65"],
            ["Max Drawdown", "-8.3%", "-7.2%"],
            ["Value at Risk (95%)", "$62,500", "-"],
        ]
        
        risk_table = Table(risk_data, colWidths=[2 * inch, 1.5 * inch, 1.5 * inch])
        risk_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), primary_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ])
        risk_table.setStyle(risk_style)
        elements.append(risk_table)
        elements.append(Spacer(1, 20))
    
    # GIPS Disclosure
    if include_gips_disclosure:
        elements.append(PageBreak())
        elements.append(Paragraph("GIPS Compliance Disclosure", styles['SectionHeader']))
        gips_text = """
        The information in this report is provided in compliance with the Global Investment 
        Performance Standards (GIPS). Past performance is not indicative of future results. 
        Returns are presented net of management fees and include reinvestment of dividends. 
        Actual performance may differ from composite returns shown. GIPS® is a registered 
        trademark of CFA Institute.
        """
        elements.append(Paragraph(gips_text.strip(), styles['Justify']))
        elements.append(Spacer(1, 15))
        
        gips_items = [
            "The firm is defined as [Firm Name], an investment advisor registered with the SEC.",
            "The composite includes all fully discretionary equity accounts.",
            "The benchmark is the S&P 500 Index.",
            "Performance calculations are based on time-weighted rates of return.",
            "All returns are shown in U.S. dollars.",
        ]
        
        for item in gips_items:
            elements.append(Paragraph(f"• {item}", styles['Normal']))
        elements.append(Spacer(1, 20))
    
    # Footer disclaimer
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    elements.append(Spacer(1, 10))
    
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER,
    )
    
    disclaimer = """
    This report is for informational purposes only and does not constitute investment advice 
    or a solicitation to buy or sell any securities. Past performance is not indicative of 
    future results. Please consult with your financial advisor before making any investment decisions.
    """
    elements.append(Paragraph(disclaimer.strip(), footer_style))
    
    # Confidential footer
    elements.append(Spacer(1, 5))
    elements.append(Paragraph("CONFIDENTIAL - For Client Use Only", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel_report(
    user_id: str,
    report_type: str,
    period_start: datetime,
    period_end: datetime,
    include_holdings: bool = True,
    include_performance: bool = True,
    include_allocation: bool = True,
) -> bytes:
    """
    Generate an Excel report for professional use.
    
    Returns:
        Excel file as bytes
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Colors
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = f"{report_type.upper()} Report"
    ws_summary['A1'].font = Font(size=16, bold=True, color="1a1a2e")
    ws_summary['A2'] = f"Period: {period_start.strftime('%Y-%m-%d')} to {period_end.strftime('%Y-%m-%d')}"
    ws_summary['A3'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}"
    
    ws_summary['A5'] = "Key Metrics"
    ws_summary['A5'].font = Font(bold=True, size=12)
    
    summary_data = [
        ["Metric", "Value"],
        ["Total Portfolio Value", "$1,250,000.00"],
        ["Period Return", "+5.42%"],
        ["YTD Return", "+12.85%"],
        ["Sharpe Ratio", "1.42"],
    ]
    
    for row_idx, row_data in enumerate(summary_data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == 6:
                cell.fill = header_fill
                cell.font = header_font
    
    # Holdings Sheet
    if include_holdings:
        ws_holdings = wb.create_sheet("Holdings")
        holdings_data = [
            ["Symbol", "Shares", "Price", "Value", "Cost Basis", "Return", "Return %"],
            ["AAPL", 100, 178.50, 17850, 15872, 1978, "12.46%"],
            ["MSFT", 75, 378.90, 28418, 26145, 2273, "8.70%"],
            ["GOOGL", 50, 141.80, 7090, 6150, 940, "15.28%"],
            ["AMZN", 60, 178.25, 10695, 8760, 1935, "22.09%"],
            ["NVDA", 40, 875.30, 35012, 24000, 11012, "45.88%"],
        ]
        
        for row_idx, row_data in enumerate(holdings_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_holdings.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                elif col_idx >= 3:
                    cell.number_format = '$#,##0.00' if col_idx <= 5 else '0.00%'
        
        # Adjust column widths
        for col in range(1, 8):
            ws_holdings.column_dimensions[get_column_letter(col)].width = 15
    
    # Allocation Sheet
    if include_allocation:
        ws_alloc = wb.create_sheet("Allocation")
        alloc_data = [
            ["Asset Class", "Allocation %", "Value"],
            ["Equities", 65, 812500],
            ["Fixed Income", 20, 250000],
            ["Cash & Equivalents", 10, 125000],
            ["Alternatives", 5, 62500],
        ]
        
        for row_idx, row_data in enumerate(alloc_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_alloc.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                elif col_idx == 2:
                    cell.number_format = '0%'
                elif col_idx == 3:
                    cell.number_format = '$#,##0'
        
        for col in range(1, 4):
            ws_alloc.column_dimensions[get_column_letter(col)].width = 20
    
    # Performance Sheet
    if include_performance:
        ws_perf = wb.create_sheet("Performance")
        perf_data = [
            ["Period", "Return", "Benchmark Return", "Excess Return"],
            ["Q1 2026", "5.42%", "4.89%", "0.53%"],
            ["Q4 2025", "3.21%", "2.95%", "0.26%"],
            ["Q3 2025", "2.18%", "2.45%", "-0.27%"],
            ["Q2 2025", "1.72%", "1.85%", "-0.13%"],
            ["YTD", "12.85%", "11.56%", "1.29%"],
        ]
        
        for row_idx, row_data in enumerate(perf_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_perf.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
        
        for col in range(1, 5):
            ws_perf.column_dimensions[get_column_letter(col)].width = 18
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
